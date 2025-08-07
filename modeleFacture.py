from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image

def saisir_facture(raison_social, produit_analyses):

    wb = Workbook()
    ws = wb.active

    # Margins
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0
    ws.page_margins.right = 0
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    # Column widths
    col_widths = {'A': 12, 'B': 25, 'E': 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Image
    img = Image("image.png")
    img.width = 100
    img.height = 75
    ws.add_image(img, "A2")

    # Styles
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border_top = Border(top=Side(style='thin'))
    border_bottom = Border(bottom=Side(style='thin'))
    border_left = Border(left=Side(style='thin'))
    border_right = Border(right=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    titre_font = Font(name='Calibri', size=10, bold=True)
    acssqda_info_font = Font(name='Calibri', size=7, italic=True)
    normal_font = Font(name='Calibri', size=8, italic=True)

    # Cell definitions: (cell, value, font, border, alignment)
    cells = [
        # Colonne A
        ('A7', "NIF:", acssqda_info_font, None, None),
        ('A8', "STAT:", acssqda_info_font, None, None),
        ('A9', "TEL:", acssqda_info_font, None, None),
        ('A10', "Adresse:", acssqda_info_font, None, None),
        ('A13', "Numéro", normal_font, Border(top=Side(style='thin'), left=Side(style='thin')), None),
        ('A14', "", normal_font, border_left, None),
        ('A15', "Réf.", normal_font, Border(top=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin')), None),
        ('A16', "Bulletin", normal_font, Border(right=Side(style='thin'), left=Side(style='thin')), None),
        ('A17', "d'analyse", normal_font, Border(bottom=Side(style='thin'), right=Side(style='thin'), left=Side(style='thin')), None),
        ('A44', "Arrêtée la présente facture à la somme de : Ariary ", normal_font, None, None),
        ('A46', "Mode de paiement", normal_font, None, None),
        ('A50', "(*) Chèque vise à l'ordre de Madame le RECEVEUR GENERAL .", normal_font, None, None),
        # Colonne B
        ('B2', "AGENCE DE CONTRÔLE DE", titre_font, border_top, align_center),
        ('B3', "LA SECURITE SANITAIRE", titre_font, None, align_center),
        ('B4', "ET DE LA QUALITE DES", titre_font, None, align_center),
        ('B5', "DENREES ALIMENTAIRES", titre_font, border_bottom, align_center),
        ('B7', "2001451249", acssqda_info_font, None, None),
        ('B8', "86,909,112,006,001,800", acssqda_info_font, None, None),
        ('B9', "22 222 39", acssqda_info_font, None, None),
        ('B10', "Rue Karidja Tsaralalàna", acssqda_info_font, None, None),
        ('B11', "(Ex Bâtiment Pharmacie Centrale Face Hôtel de Police)", acssqda_info_font, None, None),
        ('B13', "Date d'émission", normal_font, border_top, None),
        ('B15', "", None, Border(top=Side(style='thin'), right=Side(style='thin')), None),
        ('B16', "Désignations", normal_font, border_right, None),
        ('B17', "", None, Border(bottom=Side(style='thin'), right=Side(style='thin')), None),
        ('B47', "", normal_font, None, None),
        # Colonne C
        ('C12', "FACTURE", Font(name='Calibri', size=12, bold=True, italic=True), None, None),
        ('C13', "Référence(s) des produits", normal_font, border_top, None),
        ('C15', "", None, Border(top=Side('thin'), right=Side('thin')), None),
        ('C16', "N°Acte de", normal_font, border_right, None),
        ('C17', "prélèvement", normal_font, Border(bottom=Side('thin'), right=Side('thin')), None),
        ('C46', "Espèces", normal_font, None, None),
        # Colonne D
        ('D2', "Raison social:", normal_font, Border(top=Side('thin'), left=Side('thin')), None),
        ('D3', "", None, border_left, None),
        ('D4', "Statistique:", normal_font, border_left, None),
        ('D5', "", None, border_left, None),
        ('D6', "NIF:", normal_font, border_left, None),
        ('D7', "", None, border_left, None),
        ('D8', "Adresse:", normal_font, Border(left=Side('thin'), bottom=Side('thin')), None),
        ('D15', "", None, Border(top=Side('thin'), right=Side('thin')), None),
        ('D16', "Physico", normal_font, border_right, None),
        ('D17', "chimique", normal_font, Border(right=Side('thin'), bottom=Side('thin')), None),
        ('D42', "Montant à payer", normal_font, None, None),
        ('D47', "Le Client", normal_font, None, None),
        # Colonne E
        ('E1', "DOIT", titre_font, border_bottom, None),
        ('E2', raison_social, normal_font, None, None),
        ('E8', "", None, border_bottom, None),
        ('E13', "Date du résultat", normal_font, border_top, None),
        ('E15', "", None, Border(top=Side('thin'), right=Side('thin')), None),
        ('E16', "Micro-", normal_font, border_right, None),
        ('E17', "biologique", normal_font, Border(bottom=Side('thin'), right=Side('thin')), None),
        ('E46', "Chèque", normal_font, None, None),
        # Colonne F
        ('F1', "", None, border_bottom, None),
        ('F8', "", None, border_bottom, None),
        ('F13', "Responsable", normal_font, border_top, align_center),
        ('F15', "", None, Border(top=Side('thin'), right=Side('thin')), None),
        ('F16', "Toxico-", normal_font, border_right, None),
        ('F17', "logique", normal_font, Border(right=Side('thin'), bottom=Side('thin')), None),
        ('F47', "Le(a) Caissièr(e)", normal_font, None, None),
        ('F50', "Quittance N°", normal_font, None, None),
        # Colonne G
        ('G2', "", None, Border(top=Side('thin'), right=Side('thin')), None),
        ('G3', "", None, border_right, None),
        ('G4', "", None, border_right, None),
        ('G5', "", None, border_right, None),
        ('G6', "", None, border_right, None),
        ('G7', "", None, border_right, None),
        ('G8', "", None, Border(bottom=Side('thin'), right=Side('thin')), None),
        ('G13', "", None, border_right, None),
        ('G14', "", None, border_right, None),
        ('G15', "", None, Border(top=Side('thin'), right=Side('thin')), None),
        ('G16', "Sous-total", normal_font, border_right, None),
        ('G17', "", None, Border(bottom=Side('thin'), right=Side('thin')), None),
        ('G42', "Ar", normal_font, None, Alignment(horizontal='right')),
    ]

    # Set cell values, fonts, borders, alignments
    for cell, value, font, border, alignment in cells:
        ws[cell] = value
        if font:
            ws[cell].font = font
        if border:
            ws[cell].border = border
        if alignment:
            ws[cell].alignment = alignment

    # Merge cells
    merge_ranges = ['C13:D13', 'F13:G13']
    for rng in merge_ranges:
        ws.merge_cells(rng)

    # Borders for ranges
    def set_range_border(col, start, end, border):
        for row in range(start, end+1):
            ws[f"{col}{row}"].border = border

    for col in 'ABCDEFG':
        set_range_border(col, 18, 40, Border(left=Side('thin'), right=Side('thin')))
        ws[f"{col}41"].border = Border(bottom=Side('thin'), right=Side('thin'), left=Side('thin'))

    # Specific bottom borders
    for col in 'ABCDEFG':
        for row in [14, 8, 1]:
            if f"{col}{row}" in ws:
                ws[f"{col}{row}"].border = border_bottom
    
    # Entrer les désignations dans la colonne B
    row_start = 18
    for produit_analyse in produit_analyses:
        
        ws[f"A{row_start}"] = produit_analyse[1]
        ws[f"A{row_start}"].font = normal_font
        
        ws[f"B{row_start}"] = produit_analyse[0]
        ws[f"B{row_start}"].font = normal_font
        
        ws[f"C{row_start}"] = produit_analyse[2]
        ws[f"C{row_start}"].font = normal_font
        
        ws[f"D{row_start}"] = produit_analyse[3]
        ws[f"D{row_start}"].font = normal_font
        
        ws[f"E{row_start}"] = produit_analyse[4]
        ws[f"E{row_start}"].font = normal_font
        
        ws[f"F{row_start}"] = produit_analyse[5]
        ws[f"F{row_start}"].font = normal_font
        
        ws[f"G{row_start}"] = produit_analyse[6]
        ws[f"G{row_start}"].font = normal_font
        
        row_start += 1

    wb.save("modele_facture.xlsx")
    